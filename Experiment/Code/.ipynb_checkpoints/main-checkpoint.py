import os
import openpyxl
from typing import List, Tuple
from openai import OpenAI


def request(query: str, model: str) -> str:
    client = OpenAI(api_key="sk-zzzbycfqtwviaeueutrrgdjpntztnirkerttxwivwkoyend", base_url="https://api.aigc2d.com/v1")
    response = client.chat.completions.create(
        model=model,
        messages=[
            {'role': 'user', 'content': query}
        ],
        stream=False
    )
    return response.choices[0].message.content


def request_gpt(query: str) -> str:
    return request(query, 'gpt-3.5-turbo')


def request_claude(query: str) -> str:
    return request(query, 'claude-2')


def request_wenxinyiyan(query: str) -> str:
    return request(query, 'ernie-bot')


def process_excel_files(directory: str):
    # 获取目录下所有的 xlsx 文件
    excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

    for file in excel_files:
        file_path = os.path.join(directory, file)
        print(f"正在处理文件: {file}")

        try:
            # 打开 Excel 文件
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # 获取表头
            headers = [cell.value for cell in sheet[1]]

            # 获取列索引
            col_indices = get_column_indices(headers)

            # 从第二行开始处理数据（跳过表头）
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                process_row(row_index, row, sheet, col_indices, file)

                workbook.save(file_path)

            print(f"文件 {file} 处理完成")

            print(f"文件 {file} 处理完成")

        except Exception as e:
            print(f"处理文件 {file} 时发生错误: {str(e)}")
            # 发生错误时中止执行
            break


def get_column_indices(headers: List[str]) -> Tuple[int, int, int, int]:
    # 获取各列的索引
    测试用例_index = headers.index('测试用例')
    GPT_index = headers.index('GPT-3.5')
    Claude_index = headers.index('Claude')
    文心一言_index = headers.index('文心一言')

    return 测试用例_index, GPT_index, Claude_index, 文心一言_index


def process_row(row_index: int, row: Tuple, sheet: openpyxl.worksheet.worksheet.Worksheet,
                col_indices: Tuple[int, int, int, int], file: str):
    测试用例_index, GPT_index, Claude_index, 文心一言_index = col_indices

    测试用例 = row[测试用例_index]

    print(f"正在处理文件 {file} 的第 {row_index} 行")
    print(f"测试用例: {测试用例}")

    # 处理 GPT-3.5
    if not row[GPT_index]:
        try:
            print("正在请求 GPT-3.5 API...")
            response = request_gpt(测试用例)
            sheet.cell(row=row_index, column=GPT_index + 1, value=response)
            print("GPT-3.5 响应已写入")
        except Exception as e:
            print(f"请求 GPT-3.5 API 失败: {str(e)}")
            raise
    else:
        print("GPT-3.5 单元格已有内容，跳过")

    # 处理 Claude
    if not row[Claude_index]:
        try:
            print("正在请求 Claude API...")
            response = request_claude(测试用例)
            sheet.cell(row=row_index, column=Claude_index + 1, value=response)
            print("Claude 响应已写入")
        except Exception as e:
            print(f"请求 Claude API 失败: {str(e)}")
            raise
    else:
        print("Claude 单元格已有内容，跳过")

    # 处理文心一言
    if not row[文心一言_index]:
        try:
            print("正在请求文心一言 API...")
            response = request_wenxinyiyan(测试用例)
            sheet.cell(row=row_index, column=文心一言_index + 1, value=response)
            print("文心一言响应已写入")
        except Exception as e:
            print(f"请求文心一言 API 失败: {str(e)}")
            raise
    else:
        print("文心一言单元格已有内容，跳过")

    print(f"文件 {file} 的第 {row_index} 行处理完成")


# 主函数
if __name__ == "__main__":
    directory = "files"  # 替换为实际的目录路径
    process_excel_files(directory)